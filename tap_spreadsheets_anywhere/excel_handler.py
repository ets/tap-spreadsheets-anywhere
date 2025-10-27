import re
import openpyxl
import logging

import xlrd

LOGGER = logging.getLogger(__name__)

def generator_wrapper(reader, table_spec: dict={}) -> dict:
    skip_initial = table_spec.get("skip_initial", 0)
    _skip_count = 0
    header_row = None
    for row in reader:
        if _skip_count < skip_initial:
            LOGGER.debug("Skipped (%d/%d) row: %r", _skip_count, skip_initial, row)
            _skip_count += 1
            continue

        to_return = {}
        if header_row is None:
            header_row = row
            continue

        for index, cell in enumerate(row):
            header_cell = header_row[index]

            formatted_key = header_cell.value
            if not formatted_key:
                formatted_key = '' # default to empty string for key

            # remove non-word, non-whitespace characters
            formatted_key = re.sub(r"[^\w\s]", '', formatted_key)

            # replace whitespace with underscores
            formatted_key = re.sub(r"\s+", '_', formatted_key)

            to_return[formatted_key.lower()] = cell.value

        yield to_return

def get_legacy_row_iterator(table_spec, file_handle):
    workbook = xlrd.open_workbook(on_demand=True,file_contents=file_handle.read())
    if "worksheet_name" in table_spec:
        try:
            sheet = workbook.sheet_by_name(table_spec["worksheet_name"])
        except Exception as e:
            LOGGER.error("Unable to open specified sheet '"+table_spec["worksheet_name"]+"' - did you check the workbook's sheet name for spaces?")
            raise e
    else:
        try:
            sheet_name_list = workbook.sheet_names()
            #if one sheet
            if(workbook.nsheets == 1):
                sheet = workbook.sheet_by_name(sheet_name_list[0])
            #else picks sheet with most data found determined by number of rows
            else:
                sheet_list = workbook.sheets()
                max_row = 0
                max_name = ""
                for i in sheet_list:
                    if i.nrows > max_row:
                        max_row = i.nrows
                        max_name = i.name
                sheet = workbook.sheet_by_name(max_name)
        except Exception as e:
            LOGGER.info(e)
            sheet = workbook.sheet_by_name(sheet_name_list[0])
    return generator_wrapper(sheet.get_rows(), table_spec)


def get_row_iterator(table_spec, file_handle):
    data_only = table_spec.get("data_only", True)
    workbook = openpyxl.load_workbook(file_handle, read_only=True, data_only=data_only)
    
    if "worksheet_name" in table_spec:
        try:
            active_sheet = workbook[table_spec["worksheet_name"]]
        except Exception as e:
            LOGGER.error("Unable to open specified sheet '"+table_spec["worksheet_name"]+"' - did you check the workbook's sheet name for spaces?")
            raise e
    else:
        try:
            worksheets = workbook.worksheets
            #if one sheet
            if(len(worksheets) == 1):
                active_sheet = worksheets[0]
            #else picks sheet with most data found determined by number of rows
            else:
                max_row = 0
                longest_sheet_index = 0
                for i, sheet in enumerate(worksheets):
                    if sheet.max_row > max_row:
                        max_row = i.max_row
                        longest_sheet_index = i
                active_sheet = worksheets[longest_sheet_index]
        except Exception as e:
            LOGGER.info(e)
            active_sheet = worksheets[0]
    return generator_wrapper(active_sheet, table_spec)
